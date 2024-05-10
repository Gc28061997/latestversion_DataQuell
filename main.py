import os
        # import time
import pandas as pd
import json
from flask import Flask, request, jsonify, render_template, session,flash, redirect,url_for,send_file,current_app
import jwt
from functools import wraps
from configparser import ConfigParser
from flask_mysqldb import MySQL
import secrets
from werkzeug.utils import secure_filename 
from flask_bootstrap import Bootstrap
from flask_wtf import FlaskForm
from werkzeug.security import generate_password_hash, check_password_hash
from flask_sqlalchemy import SQLAlchemy
from wtforms import StringField, PasswordField, BooleanField, StringField,  SubmitField
from wtforms.validators import InputRequired, Email, Length, DataRequired, Length,EqualTo
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user
from flask_migrate import Migrate
from flask_session import Session
from itsdangerous import URLSafeTimedSerializer as Serializer
import smtplib
import ssl
import datetime
from datetime import datetime
import csv
import datetime
import schedule
import threading
import time
from datetime    import datetime, timedelta
import re
from cron_descriptor import get_description
from croniter import croniter
import config
from flask import send_from_directory


from openpyxl import load_workbook

###### Driver import  ##########

# Databricks notebook source
# import json
# import pandas as pd
import numpy as np
from datetime import datetime
from datetime import datetime, timedelta
import configparser

from Validations.CsvParser import getDFfromCsv, getDFfromXlsxMerge, getDFfromXls, check_dtype, check_ruleValidation,getDFfromXlsx
from Validations.JsonParser import GetAllValueByKey, GetRules
from Validations.Utility import getUniqueValueList, list_contains

import glob
from datetime import datetime

from werkzeug.utils import secure_filename

# Assuming this is where the error occurs
# Assuming data is obtained from a database query or some other source



#######################


app = Flask(__name__)
app.config["DEBUG"] = True
app.config['SQLALCHEMY_DATABASE_URI']='mysql://mcsvohpp_root10:welcome#2024@103.86.176.251/mcsvohpp_datavalidation'
app.config['SECRET_KEY'] = secrets.token_hex(16)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config["SESSION_PERMANENT"] = False
app.config["SESSION_TYPE"] = "filesystem"
app.config.from_object(config)
# license_key = app.config['LICENSE_KEY']
# decryption_data=app.config['DECRYPTION_DATA']
# decryption_key=app.config['DECRYPTION_KEY']

#app.config['SECRET_KEY']



Session(app)
Bootstrap(app) 
db = SQLAlchemy(app)
migrate = Migrate(app, db)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view= 'login'

# Intialize MySQL
mysql = MySQL(app)

basedir = os.path.abspath(os.path.dirname(__file__))

a=os.path.basename(os.path.dirname(__file__))
dirname=os.path.dirname(__file__)
UPLOAD_FOLDER = os.path.join(basedir,'/')
app.config['UPLOAD_FOLDER'] =  UPLOAD_FOLDER

# Now you can use the data variable

class User(UserMixin,db.Model):
    id = db.Column(db.Integer, primary_key =True)
    username = db.Column(db.String(100), nullable = False)
    email = db.Column(db.String(100), nullable= False)
    password = db.Column(db.String(150), nullable = False)
    role = db.Column(db.String(15))

    def __init__(self,username,email,password,role) :
        self.username=username
        self.email=email
        self.password=password
        self.role=role

    def __repr__(self):
            return '<User %r>' % self.username

    def get_token(self):
        serial=Serializer(app.config['SECRET_KEY'])
        return serial.dumps({'id':self.id}).encode().decode ('utf-8') 
    
    @staticmethod
    def verify_token(token):
        serial=Serializer(app.config['SECRET_KEY'])

        try:
            id=serial.loads(token)['id']
            print("id",id)
        except:
             return None
        return User.query.get(id)         
 
#with app.app_context():
#     db.create_all()
#     db.session.add(User('admin', 'admin@example.com','12345','user'))
#     db.session.add(User('guest', 'guest@example.com','12345','user'))
#     db.session.add(User('yash123', 'yash100chouhan@gmail.com','12345','admin'))
#     db.session.commit()

@login_manager.user_loader
def load_user(user_id):
	return User.query.get(user_id)

class ResetRequestForm(FlaskForm):
    email= StringField('email', validators=[InputRequired(), Length(min=5, max=45)])
    reset = SubmitField('reset')

class ResetPasswordForm(FlaskForm):
    password = PasswordField('password', validators=[InputRequired(),Length(min=5, max=10) ,EqualTo('confirm_password', message='Passwords must match')])
    confirm_password = PasswordField('confirm_password', validators=[InputRequired(), Length(min=5, max=10)])
    submit = SubmitField('submit') 

class LoginForm(FlaskForm):
	username= StringField('username', validators=[InputRequired(), Length(min=5, max=15)])
	password = PasswordField('password', validators=[InputRequired(), Length(min=5, max=10)])
	remember = BooleanField('Remember me')


class RegisterationForm(FlaskForm):
	email= StringField('Email', validators=[InputRequired(),Email(message='Invalid email'), Length(max=50)])
	username = StringField('username', validators=[InputRequired(), Length(min=4, max=15)])
	password = PasswordField('password', validators=[InputRequired(), Length(min=5, max=15)])

class UpdateForm(FlaskForm):
   
    email = StringField('email',Email(message=('Not a valid email address.')),[DataRequired()])
    username = StringField('username',[DataRequired()])
    submit = SubmitField('Submit')
     



def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        id=session.get("id")
        token=session.get("token")
        
        if not token:
            return jsonify({'message' : 'Unauthorized Access'}), 401
        if token:
            try:
                current_user = User.query.filter_by(id=id).first()
                
            except:
                return jsonify({'message': 'Something is missing in token'}), 401

            return f(current_user, *args, **kwargs)
    return decorated 


def send_mail(user):
    token=user.get_token()
    smtp_port = 587                 
    smtp_server = "smtp.gmail.com" 
    email_from = "gargi.chaurasia@zingmind.com"
    email_to = user.email
    pswd = "qfgcupcdjotabklg"
    
    message = f'''To reset ur password click on link
           {url_for('reset_token',token=token,_external=True)}
            IF YOU DID'NT SEND A PASSWORD RESET REQUEST. PLEASE IGNORE THIS MESSAGE
   '''
    simple_email_context = ssl.create_default_context()
    try:
  
        print("Connecting to server...")
        TIE_server = smtplib.SMTP(smtp_server, smtp_port)
        TIE_server.starttls(context=simple_email_context)
        TIE_server.login(email_from, pswd)
        print("Connected to server :-)")
        print()
        print(f"Sending email to - {email_to}")
        TIE_server.sendmail(email_from, email_to, message)
        print(f"Email successfully sent to - {email_to}")
    except Exception as e:
        print(e)
    finally:
     TIE_server.quit()



@app.route('/reset_password',methods=['GET','POST'])
def reset_request():
    form=ResetRequestForm()
    if form.validate_on_submit():
        user=User.query.filter_by(email=form.email.data).first()
        if user:
            send_mail(user)
            print("message sent successfully!!!")
            flash('Reset request sent. Check your mail. ','success')
            return redirect(url_for('login'))
        else:
            flash('This email is not registered. please enter registered email','danger') 
            return redirect(url_for('reset_request'))       
    return render_template('reset_request.html',title='reset request',form=form)


@app.route('/reset_password/<token>',methods=['GET','POST'])
def reset_token(token):
    user= User.verify_token(token)
    if user is None:
        flash('That is invalid token or expired','warning')
        return redirect(url_for('reset_request'))
    form=ResetPasswordForm()
    if form.validate_on_submit():
        hashed_password = generate_password_hash(form.password.data, method= 'sha256')
        user.password=hashed_password
        db.session.commit()
        print("password changed")
        flash('password changed! please login!','success')
        return redirect(url_for('login'))
       

    return render_template('change_password.html',form=form)



@app.route('/', methods=['POST'])
def verify_key():
    # input_key = request.form['key']
    # if license_key:
    #    now = datetime.datetime.now()
    #    date_format = "%Y-%m-%d"
    #    expiration_date = datetime.datetime.strptime(decryption_data["expiry_date"], date_format)
       

    #    if now > expiration_date:
    #         print("License has expired")
    #         flash('License has expired','danger')
    #         return render_template("index.html",key_valid=True)
    #    else:
            return redirect(url_for('login'))
    
        

@app.route('/',methods=['GET','POST'])
def index():
    verify_key()
    return render_template("index.html", key_valid=False)


@app.route('/editpage.html')
def edit_page():
    return render_template('editpage.html')


# Route to handle POST request for updating task record
@app.route('/updateTaskRecord', methods=['POST'])
def update_task_record():
    # Get task name and schedule from the request data
    task_name = request.form.get('taskName')  # Here, task name is fetched
    schedule = request.form.get('schedule')
    
    # Update task record in CSV file
    updated = update_task_in_csv(task_name, schedule)  # Task name is passed here
    
    if updated:
        return jsonify({'message': 'Task updated successfully'}), 200
    else:
        return jsonify({'error': 'Failed to update task'}), 500

# Function to update task record in CSV file
def update_task_in_csv(task_name, new_schedule):
    updated = False
    try:
        # Read existing data from CSV and update the record
        with open('storage.csv', mode='r', newline='') as csvfile:
            reader = csv.DictReader(csvfile)
            rows = list(reader)
        
        for row in rows:
            if row['filename'] == task_name:
                row['Schedule'] = new_schedule
        
        # Write updated data back to CSV
        with open('storage.csv', mode='w', newline='') as csvfile:
            fieldnames = ['filename', 'Last_Execution_Time', 'Overall_Result', 'Schedule', 'Human_Readable_Schedule']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
        
        updated = True
    except Exception as e:
        print(f'Error updating task: {e}')
    
    return updated


# Route to handle POST request for removing task record
@app.route('/removeTaskRecord', methods=['POST'])
def remove_task_record():
    # Get task name from the request data
    task_name = request.form.get('taskName')

    # Remove task record from CSV file
    removed = remove_task_from_csv(task_name)

    if removed:
        return jsonify({'message': 'Task removed successfully'}), 200
    else:
        return jsonify({'error': 'Failed to remove task'}), 500

# Function to remove task record from CSV file
def remove_task_from_csv(task_name):
    removed = False
    try:
        # Read existing data from CSV and remove the record
        with open('storage.csv', mode='r', newline='') as csvfile:
            reader = csv.DictReader(csvfile)
            rows = [row for row in reader if row['Task Name'] != task_name]

        # Write updated data back to CSV
        with open('storage.csv', mode='w', newline='') as csvfile:
            fieldnames = ['Task Name', 'Last_Execution_Time', 'Overall_Result', 'Schedule']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)

        removed = True
    except Exception as e:
        print(f'Error removing task: {e}')

    return removed










CSV_FILE_PATH="C:\rulengine_master\storage.csv"
def read_csv_data(csv_file_path):
    data = []
    with open(csv_file_path, 'r', newline='') as csv_file:
        csv_reader = csv.DictReader(csv_file)
        for row in csv_reader:
            # Check if Human_Readable_Schedule already exists
            if 'Human_Readable_Schedule' not in row:
                # Convert cron expression to human-readable format
                human_readable_schedule = convert_cron_to_human_readable(row['Schedule'])
                # Add the human-readable schedule to the row
                row['Human_Readable_Schedule'] = human_readable_schedule
            data.append(row)
    return data

def write_csv_data(csv_file_path, data):
    with open(csv_file_path, 'w', newline='') as csv_file:
        fieldnames = data[0].keys() if data else []
        csv_writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        csv_writer.writeheader()
        csv_writer.writerows(data)

@app.route('/loadCSVData')
def load_csv_data():
    data = read_csv_data(CSV_FILE_PATH)
    return jsonify(data)

@app.route('/updateTask', methods=['POST'])
def update_task_in_csv_route():
    # Get task name and new schedule from the request data
    task_name = request.form.get('taskName')
    new_schedule = request.form.get('newSchedule')
    
    # Update task record in CSV file
    updated = update_task_in_csv(task_name, new_schedule)
    
    if updated:
        return jsonify({'message': 'Task updated successfully'}), 200
    else:
        return jsonify({'error': 'Failed to update task'}), 500

# def convert_cron_to_human_readable(cron_expression):
#     try:
#         description = get_description(cron_expression)
#         return description
#     except Exception as e:
#         print("Error:", e)
#         return "Invalid cron expression"

# def extract_filename(rule_name):
#     filename = rule_name.replace('.csv', '').rsplit('_', 1)[0] + '.json'
#     return filename

# def append_file_to_csv(csv_file_path, rule_name, schedule):
#     try:
#         current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
#         overall_result = "Pass"
#         filename = extract_filename(rule_name)
        
#         print("Attempting to append filename:", filename)

#         # Read existing data from CSV
#         data = read_csv_data(csv_file_path)

#         # Check if filename already exists in the CSV data
#         filename_exists = any(row['Filename'] == filename for row in data)

#         if not filename_exists:
#             new_row = {'Filename': filename, 'Last Execution Time': current_time, 'Overall Result': overall_result, 'Schedule': schedule}
#             with open(csv_file_path, 'a', newline='') as csv_file:
#                 fieldnames = ['Filename', 'Last Execution Time', 'Overall Result', 'Schedule']
#                 writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
#                 writer.writerow(new_row)
#         else:
#             print("Filename", filename, "already exists in CSV file. Skipping append.")
#     except Exception as e:
#         print("An error occurred while appending to CSV file:", e)




    



# @app.route('/dashboard')
# def dashboard():
#     try:
#         csv_file_path = "storage.csv"
#         user_information_csv_path = "information.csv"

#         rule_name = None
#         schedule = None
#         with open(user_information_csv_path, 'r', newline='') as user_info_file:
#             csv_reader = csv.DictReader(user_info_file)
#             for row in csv_reader:
#                 rule_name = row.get('RuleName')
#                 schedule = row.get('Schedule')
#                 break

#         if rule_name:
#             append_file_to_csv(csv_file_path, rule_name, schedule)

#         data = read_csv_data(csv_file_path)
#         return render_template('dashboard.html', data=data)
#     except Exception as e:
#         print("An error occurred in dashboard route:", e)
#         return render_template('error.html')

# def convert_cron_to_human_readable(cron_expression):
#     try:
#         description = get_description(cron_expression)
#         return description
#     except Exception as e:
#         print("Error:", e)
#         return "Invalid cron expression"


def convert_cron_to_human_readable(cron_expression):
    try:
        description = get_description(cron_expression)
        return description
    except Exception as e:
        print("Error:", e)
        return "Invalid cron expression"


def extract_filename(rule_name):
    filename = rule_name.replace('.csv', '').rsplit('_', 1)[0] + '.json'
    return filename

def append_file_to_csv(csv_file_path, rule_name, schedule, human_readable_cron):
    try:
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        overall_result = "Pass"
        filename = extract_filename(rule_name)
        
        print("Attempting to append filename:", filename)

        # Read existing data from CSV
        data = read_csv_data(csv_file_path)

        # Check if filename already exists in the CSV data
        filename_exists = any(row['Filename'] == filename for row in data)

        if not filename_exists:
            new_row = {'Filename': filename, 'Last Execution Time': current_time, 'Overall Result': overall_result, 'Schedule': schedule, 'Human Readable Cron': human_readable_cron}
            with open(csv_file_path, 'a', newline='') as csv_file:
                fieldnames = ['Filename', 'Last Execution Time', 'Overall Result', 'Schedule', 'Human Readable Cron']
                writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
                writer.writerow(new_row)
        else:
            print("Filename", filename, "already exists in CSV file. Skipping append.")
    except Exception as e:
        print("An error occurred while appending to CSV file:", e)

@app.route('/dashboard')
def dashboard():
    try:
        csv_file_path = "storage.csv"
        user_information_csv_path = "information.csv"

        rule_name = None
        schedule = None
        human_readable_cron = None
        with open(user_information_csv_path, 'r', newline='') as user_info_file:
            csv_reader = csv.DictReader(user_info_file)
            for row in csv_reader:
                rule_name = row.get('RuleName')
                schedule = row.get('Schedule')
                human_readable_cron = convert_cron_to_human_readable(schedule)
                break

        if rule_name:
            append_file_to_csv(csv_file_path, rule_name, schedule, human_readable_cron)

        data = read_csv_data(csv_file_path)
        return render_template('dashboard.html', data=data)
    except Exception as e:
        print("An error occurred in dashboard route:", e)
        return render_template('error.html')


  


###################################################################################



# # # Function to run the scheduler


def validate_cron_expression(cron_expression):
    # Regular expression pattern to match a valid cron expression
    cron_pattern = r'^(\*|[0-5]?\d|\*\/\d+) (\*|[0-1]?\d|2[0-3]|\*\/\d+) (\*|[1-9]|[12]\d|3[01]|\*\/\d+) (\*|[1-9]|1[0-2]|\*\/\d+) (\*|[0-6]|\*\/\d+)$'
    if not re.match(cron_pattern, cron_expression):
        return False, "Invalid cron expression format"
    
    # Split the cron expression to validate the time part
    parts = cron_expression.split()
    time_part = parts[1] + ":" + parts[0]  # Concatenate hour and minute
    
    # Validate the time format
    if not validate_time_format(time_part):
        return False, "Invalid time format"
    
    return True, "Valid cron expression and time format"


def validate_time_format(time_str):
    return re.match(r'^([01]\d|2[0-3]):[0-5]\d(:[0-5]\d)?$', time_str) is not None

# Function to run the scheduler
def run_scheduler():
    while True:
        schedule.run_pending()
        time.sleep(1)  # Sleep for 1 second before checking again







# Function to schedule a job based on cron expression
def schedule_job(selected_task, schedule_time):
    try:
        print(f"parse task before Schedule '{selected_task}' with cron expression: {schedule_time}")

        # Schedule the job using the provided cron expression
        iter = croniter(schedule_time)
        next_run_time = iter.get_next(datetime)
        schedule.every().day.at(next_run_time.strftime("%H:%M")).do(run_driver)

        # Print a message indicating that job scheduling is successful
        print(f"Task '{selected_task}' scheduled successfully for {schedule_time}")
       

    except Exception as e:
        print("Error:", e)



@app.route('/driver')
def run_driver():
    try:
        print("Executing the run_driver")
        configFilePath= "configuration.ini"
        subject = 'Data_Validation'
        parser = configparser.ConfigParser()
        parser.read(configFilePath)

        #ruleFilePath = request.get('RULE_FILE_PATH')
        ruleFilePath=parser.get('APP', 'RULE_FILE_PATH')
        Column_address=parser.get('SOURCE', 'COLUMN_ADDRESS')
        Column_address1=parser.get('SOURCE', 'COLUMN_ADDRESS1')
        sheet_name=parser.get('SOURCE', 'sheet_name')

        print(Column_address)

        start_time = datetime.now()

        # Your existing function code here
        # ...

        # Record the end time with date
        end_time = datetime.now()

        # Calculate the time taken for the function execution
        execution_time = end_time - start_time
        print("Function execution time:", execution_time)

        # Store timing information in a text file
        timing_info = f"Start Time: {start_time}\nEnd Time: {end_time}\nExecution Time: {execution_time}\n"

        with open("timing_info.txt", "w") as f:
            f.write(timing_info)
        
        print("rule file path",ruleFilePath)
        outputDir = parser.get('APP', 'OUTPUT_FILE_PATH')
        reportOutputDir = outputDir + '/report/'
        errorOutputDir = outputDir + '/error/'
        rules = GetRules(ruleFilePath)
        print("Total rules in the rule file - ", len(rules))

        date = datetime.now().strftime("%Y%m%d_%I%M%S")
        TablesName = GetAllValueByKey(rules, "DataObject")
        TableList = getUniqueValueList(TablesName)

        # Data type validatoin
        sysValidDataTypeList = ['int', 'string', 'bool', 'float']
        temp = GetAllValueByKey(rules, "DataType")
        dataTypeList = getUniqueValueList(temp)
        dataTypeValidation = list_contains(sysValidDataTypeList, dataTypeList)
        print("Data type in rule file is valid - " + str(dataTypeValidation))

        # Operator validation
        sysValOperatorList = ['regex', 'reference', 'notnull', ]
        temp = GetAllValueByKey(rules, "ValidationOperator")
        operatorList = getUniqueValueList(temp)
        operatorValidator = list_contains(sysValOperatorList, operatorList)
        print("Operator validation in rule file is valid - " + str(operatorValidator))

        SOURCE_TYPE = parser.get('APP', 'SOURCE_TYPE')
        SOURCE_DATA_FILE_PATH = parser.get('SOURCE', 'SOURCE_DATA_FILE_PATH')
        SOURCE_DATA_FILE_PATH_XLS =  parser.get('SOURCE', 'SOURCE_DATA_FILE_PATH')
        SKIP_ROWS = parser.get('SOURCE', 'SKIP_ROWS')

        csvdf = pd.DataFrame()
        # Read the CSV
        if SOURCE_TYPE == 'CSV':
            csvdf = getDFfromCsv(SOURCE_DATA_FILE_PATH, SKIP_ROWS) 
            no_of_rows,no_of_columns=csvdf.shape   


        # Read the XLS
        if SOURCE_TYPE == 'XLS':
            csvdf = getDFfromXls(SOURCE_DATA_FILE_PATH, SKIP_ROWS)
            no_of_rows,no_of_columns=csvdf.shape

        # Read the XLSX
        if SOURCE_TYPE == 'XLSX':
            #csvdf = getDFfromXlsxMerge(SOURCE_DATA_FILE_PATH, SKIP_ROWS)
            csvdf= getDFfromXlsx(SOURCE_DATA_FILE_PATH, sheet_name,Column_address,Column_address1,SKIP_ROWS)
            no_of_rows,no_of_columns=csvdf.shape
        
            

        dtp = []
        for colName in csvdf.columns.tolist():
            dtp.append(check_dtype(csvdf, colName))
        # dtp = []
        # for colName in csvdf.columns.tolist():
        #     csv_col = csvdf[colName]  # Use .loc or .iloc to get a view
        #     dtp.append(check_dtype(csv_col, colName))  # Pass the DataFrame view to the function

        ruleColList = GetAllValueByKey(rules, "DataAttribute")
        ruleColList = getUniqueValueList(ruleColList)

        csvColList = csvdf.columns
        csvColList = getUniqueValueList(csvColList)   

        csvcolValidator = list_contains(ruleColList, csvColList)
        rulcolValidator = list_contains(csvColList, ruleColList)

        df_rule = pd.DataFrame(rules)
        df_rdf = pd.DataFrame()
        df_rdf = df_rule[['RuleID', 'Sequence' ,'DataAttribute', 'DataType']]
        df_rdf = df_rdf.loc[df_rdf['Sequence'].notna()]
        df_rdf.sort_values(by=['Sequence'], ascending=False)
        df_rdf = df_rdf.drop_duplicates(['DataAttribute', 'DataType'], keep='first').reset_index(drop=True)
        df_csv = pd.DataFrame(list(zip(csvColList, dtp)),columns=['CSV_Col_Name', 'CsvDataType'])
        cols_index = [(csvdf.columns.get_loc(col)+1) for col in csvColList]
        df_rdf = df_rdf.assign(CsvSequence = cols_index)
        df_rdf['CsvSequence'] = df_rdf['CsvSequence'].astype(str)
        
        bigdata = pd.concat([df_rdf, df_csv], axis=1).reindex(df_rdf.index)
        
        bigdata['CsvDataType'] = np.where(bigdata['CsvDataType'] == 'pass', bigdata['DataType'], bigdata['CsvDataType'])
        print(bigdata['CsvDataType']) 
        bigdata['CsvDataType'] = np.where(bigdata['DataType'] == 'date', bigdata['DataType'], bigdata['CsvDataType'])

        bigdata['Data_Type_Match'] = np.where(bigdata['DataType'] == bigdata['CsvDataType'], 'True', 'False')
        
    
        bigdata['Column_Match'] = np.where(df_rdf['DataAttribute'] == bigdata['CSV_Col_Name'], 'True', 'False')
        
        bigdata['SequenceMatch'] = np.where(df_rdf['CsvSequence']== df_rdf['Sequence'], 'True', 'False')
        
        
        # Write bigdata dataframe to a csv
        Rule_Summary = pd.DataFrame(columns=['RuleId', 'RuleName',  'ColumnName', 'Validation_Result','Datatype_RuleFile','Datatype_OrignalFile','Datatype_Match','RuleSequence','SequenceMatch'])
                        

        if "False" in list(bigdata['Data_Type_Match']) :
            Data_Type_Match_Flag = "Fail"
        else:
            Data_Type_Match_Flag = "Pass"

        if "False" in list(bigdata['Column_Match']) or "False" in list(bigdata['SequenceMatch']):
            Column_Match_Flag = "Fail"
            Sequence_Match_Flag="Fail"
        else:
            Column_Match_Flag = "Pass"
            Sequence_Match_Flag="Pass"

        PreSchema_Checked = bigdata.loc[bigdata['Data_Type_Match'] == 'False'] 
        PreSchema_Checked1 = bigdata.loc[bigdata['Column_Match'] == 'False'] 
        PreSchema_Checked2 = bigdata.loc[bigdata['SequenceMatch'] == 'False'] 
    
        Rule_Summary.loc['0'] = ['NA', 'Precheck-1', 'Precheck-Schema', Column_Match_Flag,'NA','NA','NA','NA',Sequence_Match_Flag]
        Rule_Summary.loc['1'] = ['NA', 'Precheck-2', 'Precheck-Datatype', Data_Type_Match_Flag,'NA','NA',Data_Type_Match_Flag,'NA','NA']
        
        
        if "Fail" == Data_Type_Match_Flag or "Fail" == Column_Match_Flag:
            fileName = "Report_" + subject + "_" + date + ".csv"
            Rule_Summary.to_csv(reportOutputDir + fileName, index=False)
            #dbutils.notebook.exit("Primary checks failed. Stop Execution")

        errFileName = "Error_" + subject + "_" + date + ".csv"

        err_out_df = pd.DataFrame()
        
        for rule in rules:
            # if rule["ValidationOperator"] == 'None': 
            #     continue
            # else:
                #print(rule["DataAttribute"], rule["RuleID"], rule["ValidationOperator"], rule["ValueToBeMatch"])
            
                Orignal_Datatype = check_dtype(csvdf,rule["DataAttribute"])
            #if Column_Match_Flag == "Pass":
                var = check_ruleValidation(csvdf, rule["DataAttribute"], rule["RuleID"], rule["ValidationOperator"],
                                        rule["ValueToBeMatch"])
                'RuleId', 'RuleName',  'ColumnName', 'Validation_Result'
                
                if Orignal_Datatype==rule["DataType"] :
                    Datatype_Match_flag="True"
                else:
                    Datatype_Match_flag="False"
                # if df_rdf['CsvSequence'].all()== df_rdf['Sequence'].all() :
                #     SequenceMatch_flag="True"
                # else:
                #     SequenceMatch_flag="False"        
                
                
                df = { 'RuleId': rule["RuleID"], 'RuleName': rule["RuleName"],
                    'ColumnName': rule["DataAttribute"],'ValidationOperator':rule["ValidationOperator"],'Datatype_RuleFile': rule["DataType"],'Datatype_OrignalFile':Orignal_Datatype,'Datatype_Match':Datatype_Match_flag,'RuleSequence': rule["Sequence"],'SequenceMatch':'', 'Validation_Result': var}
                
                df=pd.DataFrame.from_dict(df, orient='index').transpose()
                
                
                Rule_Summary = pd.concat([Rule_Summary, df]).reset_index(drop=True)
        
        cols_index.insert(0,'NA')
        cols_index.insert(1,'NA')
        
        if len(Rule_Summary['RuleId'])==len(cols_index):
            Rule_Summary['Orignalsequence']=cols_index
            Rule_Summary['Orignalsequence'] = Rule_Summary['Orignalsequence'].astype(str)
        else:
            rule_length = len(Rule_Summary['RuleId'])
            original_length = len(cols_index)
            while original_length < rule_length:
                cols_index.append('NA')
                original_length +=1
            
                        
        Rule_Summary['Orignalsequence']=cols_index
        Rule_Summary['Orignalsequence'] = Rule_Summary['Orignalsequence'].astype(str)
            
        Rule_Summary['SequenceMatch'] = np.where((Rule_Summary['Orignalsequence'].astype(str) == Rule_Summary['RuleSequence'].astype(str)), 'True', 'False')
        Rule_Summary.loc[1]['SequenceMatch']="NA"


        if "False" in list(Rule_Summary['SequenceMatch']):
            Rule_Summary.loc[0]['SequenceMatch']="False"
            Rule_Summary.loc[0]['ValidationOperator']="NA"
            Rule_Summary.loc[1]['ValidationOperator']="NA"
            
        else:
            Rule_Summary.loc[0]['SequenceMatch']="True"
            Rule_Summary.loc[0]['ValidationOperator']="NA"
            Rule_Summary.loc[1]['ValidationOperator']="NA"
        

        
                
        html_table = PreSchema_Checked.to_html(index=False, header=True, index_names=False) 
        html_table1 = PreSchema_Checked1.to_html(index=False, header=True, index_names=False) 
        html_table2 = PreSchema_Checked2.to_html(index=False, header=True, index_names=False) 
        fileName = "Report_" + subject + "_" + date + ".csv"
        Rule_Summary.to_csv( reportOutputDir + fileName, index=False)

        

        if  Rule_Summary["Validation_Result"][1] =="Fail" or Rule_Summary["Validation_Result"][0] =="Fail":
            isRuleValidationPass="False"
            ruleValidation_dict=Rule_Summary.to_dict('records')
        else:
            isRuleValidationPass="True"
            ruleValidation_dict=Rule_Summary.to_dict('records')


        if  Rule_Summary["Datatype_Match"][1] =="Fail" :    
            isDatatypeValidationPass="False"
            datatypeValidation_dict=PreSchema_Checked.to_dict('records')
        else:
            isDatatypeValidationPass="True"
            datatypeValidation_dict=PreSchema_Checked.to_dict('records')


        if  Rule_Summary["SequenceMatch"][0] =="False" :        
            isSchemaValidationPass="False"
            schemaValidation_dict=PreSchema_Checked1.to_dict('records')
            schemaValidation_dict=PreSchema_Checked2.to_dict('records')
        else:
            isSchemaValidationPass="True"
            schemaValidation_dict=PreSchema_Checked2.to_dict('records')

        
        json_object = []

        data = {'isRuleValidationPass':isRuleValidationPass,'ruleValidation':ruleValidation_dict ,'isSchemaValidationPass':isSchemaValidationPass,'schemaValidation': schemaValidation_dict,'isDatatypeValidationPass':isDatatypeValidationPass,'DatatypeValidation':datatypeValidation_dict}
        
        json_object.append(data)
            
        with open ('report.json','w') as f:
            f.write(json.dumps(json_object,indent=4))              
        

        rs_rows,rs_columns=Rule_Summary.shape
        #Write the same report csv as html file
        fileName = "Report_" + subject + "_" + date + ".html"
        html = Rule_Summary.to_html()
        html = f"<html><body>{html}<h4>Total number of rows in Csv={no_of_rows}<br>Total number of columns in Csv={no_of_columns}</h4><br></body></html>"
        if Rule_Summary["Datatype_Match"][1] =="Fail"  :
            html = f"<html><body>{html}<br><table>{html_table}</table><br></body></html>"
        # if "Fail" == Column_Match_Flag :
        #     html = f"<html><body>{html}<br><table>{html_table1}</table><br></body></html>"
        if Rule_Summary["SequenceMatch"][0] =="False":
            html = f"<html><body>{html}<br><table>{html_table2}</table><br></body></html>"

        filePath =  reportOutputDir
        text_file = open(filePath + fileName, "w")
        text_file.write(html)
        text_file.close()

        stage_file =  "C:/rulengine_master/Report/error/error_Log.csv"
        df = pd.read_csv(stage_file)
        df.drop_duplicates().to_csv( errorOutputDir + errFileName, index=False)
        print("Rwlesummary\n",Rule_Summary)
        return render_template('driver.html',Rule_Summary=Rule_Summary)
                

    except Exception as e:
        print("Error:", e)

scheduler_thread = threading.Thread(target=run_scheduler)
scheduler_thread.start()



# Function to save a task with its cron expression to the CSV file and schedule it
def save_and_schedule_task(selected_task, schedule_time):
    if not validate_cron_expression(schedule_time):
        return "Invalid cron expression"

    # Get the current date and time
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Convert cron expression to human-readable format
    human_readable_schedule = convert_cron_to_human_readable(schedule_time)

    # Save the data to a CSV file
    try:
        with open('storage.csv', 'a', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow([selected_task, current_time, 'Pass', schedule_time, human_readable_schedule])

        # Schedule the task
        print(f"Scheduling task '{selected_task}' with cron expression: {schedule_time}")
        schedule_job(selected_task, schedule_time)

        return 'Task saved and scheduled successfully.'

    except Exception as e:
        print("An error occurred while appending to CSV file:", e)
        return "Error: Task could not be saved and scheduled"





@app.route('/saveTask', methods=['POST'])
def save_task():
    selected_task = request.form['task']
    schedule_time = request.form['schedule']

    # Save and schedule the task
    save_and_schedule_task(selected_task, schedule_time)

    return 'Task saved and scheduled successfully.'


 

@app.route("/cronexpress")
@login_required
def cronjob():
    return render_template('cronexpress.html')


@app.route('/view_details')
def view_details():
    # Path to the HTML file
    html_file_path = "C:/rulengine_master/Report/Report/Report_Data_Validation_20240319_070915.html"
    
    try:
        # Open the HTML file and read its content
        with open(html_file_path, 'r') as html_file:
            html_content = html_file.read()
    except FileNotFoundError:
        # If the file is not found, return an error message
        return "HTML file not found."

    # Render the view_details.html template with the HTML content
    return render_template('view_details.html', html_content=html_content)

@app.route('/view_json/<filename>')
def view_json_file(filename):
    json_folder = 'configurationvalidator'  # Change folder name accordingly
    filepath = os.path.join(json_folder, filename)
    
    if os.path.isfile(filepath):
        with open(filepath, 'r') as file:
            json_data = file.read()
        return json_data
    else:
        return "File not found", 404


@app.route('/login', methods=['GET', 'POST']) 
def login():
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data ).first()
        if not user:
            flash('Invalid User Name and Password','danger')       
            return render_template('login.html',form=form)
        if user:    
            if check_password_hash(user.password, form.password.data):
                    
                login_user(user, remember=form.remember.data)
                user1 = User.query.filter_by(username=form.username.data ).all()
                for data in user1:
                    if data.role =="admin":
                        session['logged_in']=True
                        # token = jwt.encode({'id' : data.id,'exp' : datetime.datetime.utcnow() + timedelta(seconds=10)},app.config['SECRET_KEY'], "HS256")
                        token = jwt.encode({'id' : data.id, 'exp' : datetime.utcnow() + timedelta(seconds=10)}, app.config['SECRET_KEY'], "HS256")

                        session['id']=data.id
                        session['token']=token
                        return redirect(f"/admindashboard")
                    else:
                        
                        session['logged_in']=True
                        # token = jwt.encode({'id' : data.id,'exp' : datetime.datetime.utcnow() + timedelta(seconds=10)},app.config['SECRET_KEY'], "HS256")
                        token = jwt.encode({'id' : data.id, 'exp' : datetime.utcnow() + timedelta(seconds=10)}, app.config['SECRET_KEY'], "HS256")

                        session['id']=data.id
                        session['token']=token
                        return redirect(f"/dashboard")
            else:
                flash('Invalid User Name and Password','danger')      
                return render_template('login.html',form=form)
    user1 = User.query.filter_by(username=form.username.data ).all()
    return render_template('login.html',form=form,user=user1)

         
@app.route('/admindashboard',methods=['GET','POST'])
@token_required
@login_required
def admindashboard(current_user):
    if current_user:
        pass
        return render_template("admindashboard.html", username=current_user.username)


@app.route('/userdashboard',methods=['GET','POST'])
@token_required
@login_required
def userdashboard(current_user):
    if current_user:
        pass
        return render_template("dashboard.html", username=current_user.username)   






# signup route
@app.route('/signup', methods =['POST','GET'])

def signup():
    form = RegisterationForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username = form.username.data).first()
        if not user:
            user = User(
                username = form.username.data,
                email = form.email.data,
                password = generate_password_hash(form.password.data, method= 'sha256'),
                role='user'
                       )
            db.session.add(user)
            db.session.commit()
            flash('Successfully registered','success')
            return render_template('signup.html', form=form)
           
        else:
          
            flash('User already exists. Please Log in','warning')
            return render_template('signup.html', form=form)

    return render_template('signup.html', form=form)

@app.route('/manageusers',methods=['GET'])
@login_required
def manageusers():
    userDetails=User.query.all() 
    return render_template('manageusers.html',userDetails=userDetails)

@app.route('/update/<int:id>',methods=['GET'])
@login_required
def updateRoute(id):
    if not id or id != 0:
        Entry = User.query.get(id)
        if Entry:
            userDetails=User.query.filter_by(id=id).all()
            return render_template('update.html', userDetails=userDetails)

    
@app.route('/update/<int:id>', methods=['POST','PUT'])
@login_required
def update(id):
    
    if not id or id != 0:
        userDetails = User.query.get(id)
        if userDetails:
            new_email = request.form.get('email')
            new_username = request.form.get('username')
            userDetails.email = new_email
            userDetails.username = new_username
            db.session.commit()
        flash(' Updated Successfully','success')
        userDetails=User.query.all()
        return render_template('manageusers.html',userDetails=userDetails)

@app.route('/delete/<int:id>')
@login_required
def delete(id):
    if not id or id != 0:
        userDetails = User.query.get(id)
        if userDetails:
            db.session.delete(userDetails)
            db.session.commit()
        flash('Successfully Deleted','success')
        userDetails=User.query.all()
        return render_template('manageusers.html',userDetails=userDetails)

@app.route('/logout') 
def logout():
	logout_user()
	return redirect(url_for('index'))

################## Data Validation source selection ######################

@app.route('/Admin_data_validation',methods=['POST','GET'])
@login_required
def Admin_data_validation():
        if request.form["Submitbutton"]=='SingleDataSource':
            return render_template('AdminSingleDataSource.html')
        else:
            return render_template('AdminDoubleDataSource.html')
    
############ Admin Single Data Source Validation  ##########################


def getDFfromXlsx(xlsxPath, sheet_name,Column_address,Column_address1,SKIP_ROWS):
    flag = os.path.exists(xlsxPath)
    if flag == True:
        # df = pd.read_excel(xlsxPath, engine='openpyxl', skiprows=skip_rows, dtype=object)
        # df = df.iloc[2:8,3:6]
        # print("df",df)
        wb = load_workbook(filename=xlsxPath, read_only=True)

        ws = wb[sheet_name]
        data_rows = []

        if ((Column_address == '' and Column_address1 == '') or (Column_address == Column_address and Column_address1 == '') or (Column_address == '' and Column_address1 == Column_address1) or (Column_address == '0' and Column_address1 == '0')):
            for row in ws:
                data_cols = []
                for cell in row:
                    data_cols.append(cell.value)
                data_rows.append(data_cols)
        
        else:
             for row in ws[Column_address:Column_address1]:
                data_cols = []
                for cell in row:
                    data_cols.append(cell.value)
                data_rows.append(data_cols)

        df = pd.DataFrame(data_rows[1:], columns=data_rows[0])
        df.columns = df.columns.astype(str)
        #print(df)
    #     isHeaderOn = "true"
    #     isInferSchemaOn = "false"
    #     #sheet address in excelpip install pyspark

    #     sampleAddress = "'towers'!A4"
    #     #read excelfile
    #     #df = spark.read.format("com.crealytics.spark.excel").option("header", isHeaderOn).option("inferSchema", isInferSchemaOn).option("treatEmptyValuesAsNulls", 'True').option("dataAddress", sampleAddress).load(xlsxPath)
    #     spark = SparkSession.builder.appName("ExcelRead").getOrCreate()

    #     df = spark.read.format("com.crealytics.s00park.excel")\
    #   .option("useHeader", "true")\
    #   .option("treatEmptyValuesAsNulls", "true")\
    #   .option("inferSchema", "true")\
    #   .option("addColorColumns", "False")\
    #   .load(xlsxPath)
        
        return df

@app.route("/Admin_SingleDataSource", methods=['POST','GET'])
@login_required
def Admin_SingleDataSource():
    parser = ConfigParser()
    try:
        data_source_type = request.form['datasourcetype']
        if data_source_type=='CSV':             
            file = request.files['DataSourcePath']
            filename = secure_filename(file.filename)
            #file_path=os.path.join(basedir, file.filename)
            file_path = os.path.abspath('Store_File\\'+filename)
            delimiter = request.form['Delimiter']
            output_file_path = 'C:\\rulengine_master\Report'
            SKIP_ROWS = request.form['skip_rows']
            SHEET_NAME="None"
            Column_Address="None"
            Column_Address1="None"
            data = pd.read_csv(file_path,sep=delimiter,engine='python',encoding='latin1')          
            col_list = list(data.columns)
            # print(col_list)
            data_type_list = list(data.iloc[1])

        elif data_source_type=='XLSX' or data_source_type=='XLS':             
            file = request.files['DataSourcePath']
            filename = secure_filename(file.filename)
            #file_path=os.path.join(basedir, file.filename)
            file_path = os.path.abspath('Store_File\\'+filename)
            output_file_path = 'C:\\rulengine_master\Report'
            SKIP_ROWS = request.form['skip_rows']
            SHEET_NAME=request.form['sheet_name']
            Column_Address=request.form['Column_Address']
            Column_Address1=request.form['Column_Address1']
            #data = pd.read_excel(file_path, engine='openpyxl',sheet_name=SHEET_NAME, skiprows=SKIP_ROWS, dtype=object)
            data=getDFfromXlsx(file_path, SHEET_NAME,Column_Address,Column_Address1,SKIP_ROWS)
            delimiter=','          
            col_list = list(data.columns)
                  
            
        try: 
             
            with open("C:\\rulengine_master\configuration.ini", 'w') as file:     
                file.write("")  
            parser.add_section("APP")            
            parser.set("APP",'RULE_FILE_PATH',os.getcwd()+"\\configurationvalidator\\rule_file.json")
            parser.set("APP",'SOURCE_TYPE',data_source_type)
            parser.set("APP",'OUTPUT_FILE_PATH',output_file_path)
            parser.add_section("SOURCE")
            parser.set("SOURCE",'SOURCE_DATA_FILE_PATH', file_path)
            parser.set("SOURCE",'SKIP_ROWS', SKIP_ROWS) 
            parser.set("SOURCE",'SHEET_NAME', SHEET_NAME)
            parser.set("SOURCE",'Column_Address', Column_Address)
            parser.set("SOURCE",'Column_Address1', Column_Address1)
           
            with open("C:\\rulengine_master\configuration.ini", 'w') as file: 
                parser.write(file)

        except:
             print(Exception)
             raise 
                                                                                            
        return render_template('rule_file_generator.html',file_path=file_path,delimiter=delimiter,data=data,file_name = filename, col_list=col_list,datatype_list=[get_datatype(data,colName) for  colName in col_list],len = len(col_list))
    except:
        print(Exception)
        raise
    

@app.route("/create", methods=['POST'])
@login_required 
def create_json():
    json_object = []
    try:
        i=1
        while True:
            Dict = {"RuleID": "" + str(i) + "",
            "RuleName": request.form[f"name{i}"] + " validation",            
            "DataAttribute": request.form[f'data_attribute{i}'],
            "DataType": request.form[f'datatype{i}'],
            "ValidationOperator": request.form[f'valop{i}'],
            "ValueToBeMatch": request.form[f'valtomatch{i}'],
            "Order": request.form[f'order{i}'],
            "DataObject":request.form['DataObject'],
            "DataSource":request.form['DataSource'],
            "Sequence":request.form[f'order{i}']
            
            }
            json_object = AddToJSON(json_object, Dict)
            i+=1
    except: 
        with open ('rule_file.json','w') as f:
            f.write(json.dumps(json_object,indent=4))    
        return render_template('download.html')


def AddToJSON(json_object, myDict):
    # Data to be written
    json_object.append(myDict)
    return json_object


def get_datatype(datafram,colName):
    try:
        if colName in datafram.columns:
            datatypes = datafram.dtypes[colName]
            if datatypes == 'object':
                return 'string'
            if datatypes == 'int64':
                return 'int' 
            if datatypes == 'float':
                return 'float'
            if datatypes == 'date':
                return 'date'
            if datatypes == 'time':
                return 'time'                       
            
    except:
        raise     
    # try:
    #     if type(col_name)==str:
    #         return 'string'
    #     if type(col_name.item())==int:
    #         return 'int'
    #     if type(col_name.item())==float:
    #         return 'float'
    #     if type(col_name.item())==time:
    #         return 'time'
    #     if type(col_name.item())==date:
    #         return 'date'
    # except:
    #     raise        


@app.route("/download")
@login_required
def download_file():
    downloaded_file="rule_file.json"
    return send_file(downloaded_file,as_attachment=True)

@app.route("/Regex")
@login_required
def Regex():
    return render_template('Regex.html')



################ Admin Double Data Source Validation   ###################################

@app.route("/Admin_DoubleDataSource", methods=['POST','GET']) 
@login_required
def Admin_DoubleDataSource():
    
    parser = ConfigParser()
    try:
        with open("C:\\rulengine_master\configuration.ini", 'w') as file:
            file.write("")  
        output_file_path = request.form['output_file_path'] 
        
        data_source_type = request.form['datasourcetype']        
        
        if data_source_type == 'CSV':

            file1 = request.files['DataSourcePath1'] 
            filename1=secure_filename(file1.filename)
            file_path1=os.path.join(basedir + '\Store_File\\' + file1.filename)
            delimiter1 = request.form['Delimiter1']
            skip_rows1 = request.form['skip_rows1']
            col_name_min = request.form['min_col']
            col_name_max = request.form['max_col']
            col_name_avg = request.form['avg_col']
            col_name_sum = request.form['sum_col']
            
            parser.add_section("APP")
            parser.set("APP",'SOURCE_TYPE',data_source_type)
            parser.set("APP",'OUTPUT_FILE',output_file_path)
            
            parser.add_section("SOURCE")
            parser.set("SOURCE","SOURCE_DATA_FILE_PATH", file_path1)
            parser.set("SOURCE","Delimiter", delimiter1)
            parser.set("SOURCE","skip_rows", skip_rows1)
            parser.set("SOURCE","col_name_min", col_name_min)
            parser.set("SOURCE","col_name_max", col_name_max)
            parser.set("SOURCE","col_name_avg", col_name_avg)        
            parser.set("SOURCE","col_name_sum", col_name_sum)


            with open("C:\\rulengine_master\configuration.ini", 'w') as file:
                parser.write(file)
            
        if data_source_type == 'JSON':
            file_path1 = request.form['DataSourcePath1'] 
            parser.add_section("APP")
            parser.set("APP",'SOURCE_TYPE',data_source_type)
            parser.set("APP",'OUTPUT_FILE',output_file_path)
            parser.add_section("SOURCE")
            parser.set("SOURCE","SOURCE_DATA_FILE_PATH", file_path1)
            
            with open("C:\\rulengine_master\configuration.ini", 'w') as file:
                parser.write(file)

        if data_source_type == 'XLSX':
            file1 = request.files['DataSourcePath1'] 
            file_path1=os.path.join(basedir + '\Store_File\\' + file1.filename)
            sheet_no1 = request.form['sheet_no1'] 
            skip_rows1 = request.form['skip_rows1']
            col_name_min = request.form['min_col']
            col_name_max = request.form['max_col']
            col_name_avg = request.form['avg_col']
            col_name_sum = request.form['sum_col'] 

            parser.add_section("APP")
            parser.set("APP",'SOURCE_TYPE',data_source_type)
            parser.set("APP",'OUTPUT_FILE',output_file_path)
            parser.add_section("SOURCE")
            parser.set("SOURCE","SOURCE_DATA_FILE_PATH", file_path1)
            parser.set("SOURCE","SHEET_NAME", sheet_no1)
            parser.set("SOURCE","SKIP_ROWS", skip_rows1)
            parser.set("SOURCE","col_name_min", col_name_min)
            parser.set("SOURCE","col_name_max", col_name_max)
            parser.set("SOURCE","col_name_avg", col_name_avg)        
            parser.set("SOURCE","col_name_sum", col_name_sum)
  
            with open("C:\\rulengine_master\configuration.ini", 'w') as file:
                parser.write(file)

        if data_source_type == 'ORACLE' or data_source_type == 'MYSQL':
            server1 = request.form['Server1'] 
            database1 = request.form['Database1'] 
            user1 = request.form['user1'] 
            password1 = request.form['password1'] 
            schema_name1 = request.form['schema_name1']            
            source_query_filter1 = request.form['source_query_filter1']
            col_name_min = request.form['min_col']
            col_name_max = request.form['max_col']
            col_name_avg = request.form['avg_col']
            col_name_sum = request.form['sum_col']

            parser.add_section("APP")
            parser.set("APP",'SOURCE_TYPE',data_source_type)
            parser.set("APP",'OUTPUT_FILE',output_file_path)

            parser.add_section("SOURCE")
            parser.set("SOURCE","SOURCE_SERVER", server1)
            parser.set("SOURCE","SOURCE_DATABASE", database1)
            parser.set("SOURCE","SOURCE_USER", user1)
            parser.set("SOURCE","SOURCE_PASSWORD", password1)
            parser.set("SOURCE","col_name_min", col_name_min)
            parser.set("SOURCE","col_name_max", col_name_max)
            parser.set("SOURCE","col_name_avg", col_name_avg)        
            parser.set("SOURCE","col_name_sum", col_name_sum)

            parser.add_section("vTurbineMasterData_Source")
            parser.set("vTurbineMasterData_Source","SCHEMA_NAME_SOURCE", schema_name1)
            parser.set("vTurbineMasterData_Source","SOURCE_QUERY_FILTER",source_query_filter1)

            with open("C:\\rulengine_master\configuration.ini", 'w') as file:
                parser.write(file)
         

        ###################### destination data set #####################################
                

        data_source_type = request.form['datadesttype']
        parser.set("APP",'DEST_TYPE',data_source_type)

        if data_source_type == 'CSV':

            file2 = request.files['datasourcepath2'] 
            filename2=secure_filename(file2.filename)
            file_path2=os.path.join(basedir + '\Store_File\\' + file2.filename)
            # delimiter2 = request.form['Delimiter2']
            skip_rows2 = request.form['skip_rows2']
            delimiter2 = request.form['delimiter2']         
                        
            parser.add_section("DEST")
            parser.set("DEST","DEST_DATA_FILE_PATH", file_path2)
            parser.set("DEST","Delimiter", delimiter2)
            parser.set("DEST","skip_rows", skip_rows2)
            with open("C:\\rulengine_master\configuration.ini", 'w') as file:
                parser.write(file)
           
        if data_source_type == 'JSON':
            file_path2 = request.form['datasourcepath2'] 

            parser.add_section("DEST")
            parser.set("DEST","DEST_DATA_FILE_PATH", file_path2)
         
            with open("C:\\rulengine_master\configuration.ini", 'w') as file:
                parser.write(file)


        if data_source_type == 'XLSX':
            file_path2 = request.form['DataSourcePath2'] 
            sheet_no2 = request.form['sheet_no2'] 
            skip_rows2 = request.form['skip_rows2']

            parser.add_section("DEST")
            parser.set("DEST","DEST_DATA_FILE_PATH", file_path2)
            parser.set("DEST","SHEET_NAME", sheet_no2)
            parser.set("DEST","SKIP_ROWS", skip_rows2)

            with open("C:\\rulengine_master\configuration.ini", 'w') as file:
                parser.write(file)


        if data_source_type == 'ORACLE' or data_source_type == 'MYSQL':
            server2 = request.form['Server2'] 
            database2 = request.form['Database2'] 
            user2 = request.form['user2'] 
            password2 = request.form['password2'] 
            schema_name2 = request.form['schema_name2']            
            source_query_filter2 = request.form['source_query_filter2'] 
        
            parser.add_section("DEST")
            parser.set("DEST","DEST_SERVER", server2)
            parser.set("DEST","DEST_DATABASE", database2)
            parser.set("DEST","DEST_USER", user2)
            parser.set("DEST","DEST_PASSWORD", password2)

            parser.add_section("vTurbineMasterData_Dest")
            parser.set("vTurbineMasterData_Dest","SCHEMA_NAME_DEST", schema_name2)
            parser.set("vTurbineMasterData_Dest","Destination_QUERY_FILTER",source_query_filter2)

            with open("C:\\rulengine_master\configuration.ini", 'w') as file:
                parser.write(file)
                
        return render_template('AdminDoubleDataSource.html') 
    except:
        print(Exception)
        raise
        
#######################   user   #################################

@app.route('/User_data_validation',methods=['POST','GET'])
@login_required
def User_data_validation():
        if request.form["Submitbutton"]=='SingleDataSource':
            return render_template('UserSingleDataSource.html')
        else:
            return render_template('UserDoubleDataSource.html')

    
# ************ Single Data Source Validation User **************************

# @app.route("/User_SingleDataSource", methods=['POST','GET'])
# @login_required
# def User_SingleDataSource():
#     parser = ConfigParser()
#     try:
#         data_source_type = request.form['datasourcetype']
#         if data_source_type=='CSV':             
#             file = request.files['DataSourcePath']
#             filename = secure_filename(file.filename)
#             #file_path=os.path.join(basedir, file.filename)
#             file_path = os.path.abspath('Store_File\\'+filename)
#             delimiter = request.form['Delimiter']
#             output_file_path = 'C:\\rulengine_master\Report'
#             SKIP_ROWS = request.form['skip_rows']
#             SHEET_NAME="None"
#             Column_Address="None"
#             Column_Address1="None"
#             data = pd.read_csv(file_path,sep=delimiter,engine='python',encoding='latin1')          
#             col_list = list(data.columns)
#             # print(col_list)
#             data_type_list = list(data.iloc[1])

            
            

#         elif data_source_type=='XLSX' or data_source_type=='XLS':             
#             file = request.files['DataSourcePath']
#             filename = secure_filename(file.filename)
#             # file_path=os.path.join(basedir, file.filename)
#             file_path = os.path.abspath('Store_File\\'+filename)
#             output_file_path = 'C:\\rulengine_master\Report'
#             SKIP_ROWS = request.form['skip_rows']
#             SHEET_NAME=request.form['sheet_name']
#             Column_Address=request.form['Column_Address']
#             Column_Address1=request.form['Column_Address1']
#             # data = pd.read_excel(file_path, engine='openpyxl',sheet_name=SHEET_NAME, skiprows=SKIP_ROWS, dtype=object)
#             data=getDFfromXlsx(file_path, SHEET_NAME,Column_Address,Column_Address1,SKIP_ROWS)
#             delimiter=','          
#             col_list = list(data.columns)
         
#         try:
             
#             with open("C:\\rulengine_master\configuration.ini", 'w') as file:     
#                 file.write("")  
#             parser.add_section("APP")            
#             parser.set("APP",'RULE_FILE_PATH',os.getcwd()+"\\rule_file.json")
#             parser.set("APP",'SOURCE_TYPE',data_source_type)
#             parser.set("APP",'OUTPUT_FILE_PATH',output_file_path)
#             parser.add_section("SOURCE")
#             parser.set("SOURCE",'SOURCE_DATA_FILE_PATH', file_path)
#             parser.set("SOURCE",'SKIP_ROWS', SKIP_ROWS) 
#             parser.set("SOURCE",'SHEET_NAME', SHEET_NAME)
#             parser.set("SOURCE",'Column_Address', Column_Address)
#             parser.set("SOURCE",'Column_Address1', Column_Address1)
           
#             with open("C:\\rulengine_master\configuration.ini", 'w') as file: 
#                 parser.write(file)

#         except:
#              print(Exception)
#              raise

#         return render_template('rule_file_generator.html',file_path=file_path,delimiter=delimiter,data=data,file_name = filename, col_list=col_list,datatype_list=[get_datatype(data,colName) for  colName in col_list],len = len(col_list))
#     except:
#         print(Exception)
#         raise


def find_most_recent_file_in_folder(folder_path):
    files = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.endswith('.csv')]
    if not files:
        return None
    most_recent_file = max(files, key=os.path.getmtime)
    return most_recent_file

def find_files_with_date(file_folder, date_str):
    files = []
    for file_name in os.listdir(file_folder):
        if file_name.endswith('.csv'):
            file_base, ext = os.path.splitext(file_name)
            if len(file_base.split('_')) > 1:
                try:
                    file_date_from_name = datetime.strptime(file_base.split('_')[-1], '%m%d%Y')
                    if file_date_from_name == date_str:
                        files.append(os.path.join(file_folder, file_name))
                except ValueError:
                    pass
    return files

def find_most_recent_file(files):
    if not files:
        return None
    most_recent_file = max(files, key=os.path.getmtime)
    return most_recent_file



# def handle_file_path(file_path):
#     file_name = ""

#     if not file_path:
#         return "Error: File path is empty.", None

#     if '*' in os.path.basename(file_path):
#         # If the file path contains a wildcard, get the parent directory
#         file_folder = os.path.dirname(file_path)
#         most_recent_file = find_most_recent_file_in_folder(file_folder)
#         if not most_recent_file:
#             return "Error: No suitable file found in the specified folder.", None
#         file_name = os.path.basename(most_recent_file)
#     elif os.path.isdir(file_path):
#         most_recent_file = find_most_recent_file_in_folder(file_path)
#         if not most_recent_file:
#             return "Error: No suitable file found in the specified folder.", None
#         file_name = os.path.basename(most_recent_file)
#     else:
#         # If the provided path is not a directory, use it as it is
#         file_name = os.path.basename(file_path)

#         # Check if the filename contains a date
#         date_str = file_name.split('_')[-1].split('.')[0]
#         try:
#             date_str = datetime.strptime(date_str, '%m%d%Y')
#             files_with_date = find_files_with_date(os.path.dirname(file_path), date_str)
#             if len(files_with_date) == 1 and file_path.endswith(str(date_str.date()) + '.csv'):
#                 most_recent_file = file_path
#             else:
#                 most_recent_file = find_most_recent_file(files_with_date)
#             if not most_recent_file:
#                 return "Error: No most recent file found.", None
#         except ValueError:
#             pass  # No date found in the filename

#     return None, most_recent_file

def handle_file_path(file_path):
    file_name = ""

    if not file_path:
        return "Error: File path is empty.", None

    if '*' in os.path.basename(file_path):
        # If the file path contains a wildcard, get the parent directory
        file_folder = os.path.dirname(file_path)
        most_recent_file = find_most_recent_file_in_folder(file_folder)
        if not most_recent_file:
            return "Error: No suitable file found in the specified folder.", None
        file_name = os.path.basename(most_recent_file)
    elif os.path.isdir(file_path):
        most_recent_file = find_most_recent_file_in_folder(file_path)
        if not most_recent_file:
            return "Error: No suitable file found in the specified folder.", None
        file_name = os.path.basename(most_recent_file)
    else:
        # Check if the filename contains a date
        file_base = os.path.basename(file_path)
        date_str = file_base.split('_')[-1].split('.')[0]
        try:
            date_str = datetime.strptime(date_str, '%m%d%Y')
            files_with_date = find_files_with_date(os.path.dirname(file_path), date_str)
            if len(files_with_date) == 1 and file_path.endswith(str(date_str.date()) + '.csv'):
    
                most_recent_file = file_path
            else:
                most_recent_file = find_most_recent_file(files_with_date)
            if not most_recent_file:
                return "Error: No most recent file found.", None
        except ValueError:
            # If no date is found in the filename, use the file path as-is without any date checking
            most_recent_file = file_path

    return None, most_recent_file



@app.route("/User_SingleDataSource", methods=['POST', 'GET'])
def User_SingleDataSource():
    parser = ConfigParser()
    try:
        file_path = request.form.get('DataSourcePath', '').strip()  # Getting the file path from the form

        if not file_path:
            return "Error: File path is empty."

        file_path = os.path.abspath(file_path)
        file_folder = os.path.dirname(file_path)

        error_message, most_recent_file = handle_file_path(file_path)
        if error_message:
            return error_message

        data_source_type = request.form.get('datasourcetype')
        if data_source_type in ['CSV', 'XLSX', 'XLS']:
            if data_source_type == 'CSV':
                data = pd.read_csv(most_recent_file, engine='python', encoding='latin1')
            else:
                data = pd.read_excel(most_recent_file)
            col_list = list(data.columns)
            delimiter = request.form.get('Delimiter', ',')  # Default delimiter is comma
            output_file_path = 'C:\\rulengine_master\\Report'  # Corrected path
            SKIP_ROWS = request.form.get('skip_rows', 0)  # Default skip rows is 0
            SHEET_NAME = "None"
            Column_Address = "None"
            Column_Address1 = "None"
            file_name = os.path.basename(most_recent_file)
            return render_template('rule_file_generator.html', file_name=file_name, file_path=file_path, delimiter=delimiter, data=data,
                                    col_list=col_list,
                                    datatype_list=[get_datatype(data, colName) for colName in col_list],
                                    len=len(col_list))

        # Handle other cases or return default response if no conditions are met
        elif data_source_type in ('XLSX', 'XLS'):             
            file = request.files['DataSourcePath']
            file_name = secure_filename(file.filename)
            file_path = os.path.abspath(os.path.join('Store_File', file_name))
            output_file_path = 'C:\\rulengine_master\\Report'  # Corrected path
            SKIP_ROWS = request.form['skip_rows']
            SHEET_NAME = request.form['sheet_name']
            Column_Address = request.form['Column_Address']
            Column_Address1 = request.form['Column_Address1']
            data = getDFfromXlsx(file_path, SHEET_NAME, Column_Address, Column_Address1, SKIP_ROWS)
            delimiter = ','          
            col_list = list(data.columns)

            # Save data to information.csv file
            csv_file_path = 'information.csv'
            with open(csv_file_path, mode='w', newline='') as file:
                fieldnames = ["RuleID", "RuleName", "DataAttribute", "DataType", "ValidationOperator", "ValueToBeMatch", "Order", "DataObject", "DataSource", "Sequence", "SaveDateTime"]
                writer = csv.DictWriter(file, fieldnames=fieldnames)
                writer.writeheader()
                for i, col in enumerate(col_list, start=1):
                    # Get current date and time
                    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    writer.writerow({
                        "RuleID": str(i),
                        "RuleName": f"{file_name}_{col} validation",
                        "DataAttribute": col,
                        "DataType": "string",  # You need to determine the data type dynamically
                        "ValidationOperator": "None",
                        "ValueToBeMatch": "",
                        "Order": str(i),
                        "DataObject": "table",
                        "DataSource": file_name,
                        "Sequence": str(i),
                        "SaveDateTime": current_time
                    })

            # Write configuration to .ini file
            config_file_path = "C:\\rulengine_master\\configuration.ini"  # Corrected path
            with open(config_file_path, 'w') as file:     
                file.write("")  
            parser.add_section("APP")            
            parser.set("APP", 'RULE_FILE_PATH', os.path.join(os.getcwd(), "rule_file.json"))  # Corrected path
            parser.set("APP", 'SOURCE_TYPE', data_source_type)
            parser.set("APP", 'OUTPUT_FILE_PATH', output_file_path)
            parser.add_section("SOURCE")
            parser.set("SOURCE", 'SOURCE_DATA_FILE_PATH', file_path)
            parser.set("SOURCE", 'SKIP_ROWS', SKIP_ROWS) 
            parser.set("SOURCE", 'SHEET_NAME', SHEET_NAME)
            parser.set("SOURCE", 'Column_Address', Column_Address)
            parser.set("SOURCE", 'Column_Address1', Column_Address1)
            with open(config_file_path, 'w') as file: 
                parser.write(file)

            return render_template('rule_file_generator.html', file_name=file_name, file_path=file_path, delimiter=delimiter, data=data, 
                                col_list=col_list, 
                                datatype_list=[get_datatype(data, colName) for colName in col_list], 
                                len=len(col_list))
        else:
            return "Unsupported data source type."

    except Exception as e:
        print("An error occurred:", e)
        return "An error occurred: " + str(e)






# @app.route("/User_SingleDataSource", methods=['POST','GET'])
# @login_require
# def User_SingleDataSource():
#     parser = ConfigParser()
#     try:
#         data_source_type = request.form['datasourcetype']
#         print("Data source type:", data_source_type)
#         if data_source_type == 'CSV': 
                        
#             file = request.files['DataSourcePath']
#             print("Uploaded file:", file.filename)
#             # filename = secure_filename(file.filename)
#             # print("Secure filename:", filename)
#             file_path=request.filepath(file_path)
#             file_path = os.path.abspath(os.path.join('C:\\rulengine_master\\Store_File', filename))
#             print("Absolute file path:", file_path)
#             delimiter = request.form['Delimiter']
#             print("Delimiter:", delimiter)
#             output_file_path = 'C:\\rulengine_master\\Report'  # Corrected path
#             SKIP_ROWS = request.form['skip_rows']
#             print("Skip rows:", SKIP_ROWS)
#             SHEET_NAME = "None"
#             Column_Address = "None"
#             Column_Address1 = "None"
#             data = pd.read_csv(file_path, sep=delimiter, engine='python', encoding='latin1')          
#             col_list = list(data.columns)
#             print("Column list:", col_list)

#         elif data_source_type == 'XLSX' or data_source_type == 'XLS':             
#             file = request.files['DataSourcePath']
#             filename = secure_filename(file.filename)
#             file_path = os.path.abspath(os.path.join('Store_File', filename))
#             output_file_path = 'C:\\rulengine_master\\Report'  # Corrected path
#             SKIP_ROWS = request.form['skip_rows']
#             SHEET_NAME = request.form['sheet_name']
#             Column_Address = request.form['Column_Address']
#             Column_Address1 = request.form['Column_Address1']
#             data = getDFfromXlsx(file_path, SHEET_NAME, Column_Address, Column_Address1, SKIP_ROWS)
#             delimiter = ','          
#             col_list = list(data.columns)

#         # Save data to information.csv file
#         csv_file_path = 'information.csv'
#         with open(csv_file_path, mode='w', newline='') as file:
#             fieldnames = ["RuleID", "RuleName", "DataAttribute", "DataType", "ValidationOperator", "ValueToBeMatch", "Order", "DataObject", "DataSource", "Sequence", "SaveDateTime"]
#             writer = csv.DictWriter(file, fieldnames=fieldnames)
#             writer.writeheader()
#             for i, col in enumerate(col_list, start=1):
#                 # Get current date and time
#                 current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
#                 writer.writerow({
#                     "RuleID": str(i),
#                     "RuleName": f"{filename}_{col} validation",
#                     "DataAttribute": col,
#                     "DataType": "string",  # You need to determine the data type dynamically
#                     "ValidationOperator": "None",
#                     "ValueToBeMatch": "",
#                     "Order": str(i),
#                     "DataObject": "table",
#                     "DataSource": filename,
#                     "Sequence": str(i),
#                     "SaveDateTime": current_time
#                 })

#         # Write configuration to .ini file
#         config_file_path = "C:\\rulengine_master\\configuration.ini"  # Corrected path
#         with open(config_file_path, 'w') as file:     
#             file.write("")  
#         parser.add_section("APP")            
#         parser.set("APP", 'RULE_FILE_PATH', os.path.join(os.getcwd(), "rule_file.json"))  # Corrected path
#         parser.set("APP", 'SOURCE_TYPE', data_source_type)
#         parser.set("APP", 'OUTPUT_FILE_PATH', output_file_path)
#         parser.add_section("SOURCE")
#         parser.set("SOURCE", 'SOURCE_DATA_FILE_PATH', file_path)
#         parser.set("SOURCE", 'SKIP_ROWS', SKIP_ROWS) 
#         parser.set("SOURCE", 'SHEET_NAME', SHEET_NAME)
#         parser.set("SOURCE", 'Column_Address', Column_Address)
#         parser.set("SOURCE", 'Column_Address1', Column_Address1)
#         with open(config_file_path, 'w') as file: 
#             parser.write(file)

#         return render_template('rule_file_generator.html', file_path=file_path, delimiter=delimiter, data=data, 
#                                file_name=filename, col_list=col_list, 
#                                datatype_list=[get_datatype(data, colName) for colName in col_list], 
#                                len=len(col_list))
#     except Exception as e:
#         print("An error occurred:", e)
#         raise



#************* Double Data Source Validation User ****************

@app.route("/User_DoubleDataSource", methods=['POST','GET']) 
@login_required
def User_DoubleDataSource():
    
    parser = ConfigParser()
    try:
        with open("C:\\rulengine_master\configuration.ini", 'w') as file:
            file.write("")  
        output_file_path = request.form['output_file_path'] 
        
        data_source_type = request.form['datasourcetype']        
        
        if data_source_type == 'CSV':

            file1 = request.files['DataSourcePath1'] 
            filename1=secure_filename(file1.filename)
            file_path1=os.path.join(basedir + '\Store_File\\' + file1.filename)
            delimiter1 = request.form['Delimiter1']
            skip_rows1 = request.form['skip_rows1']
            col_name_min = request.form['min_col']
            col_name_max = request.form['max_col']
            col_name_avg = request.form['avg_col']
            col_name_sum = request.form['sum_col'] 

            parser.add_section("APP")
            parser.set("APP",'SOURCE_TYPE',data_source_type)
            parser.set("APP",'OUTPUT_FILE',output_file_path)

            parser.add_section("SOURCE")
            parser.set("SOURCE","SOURCE_DATA_FILE_PATH", file_path1)
            parser.set("SOURCE","Delimiter", delimiter1)
            parser.set("SOURCE","skip_rows", skip_rows1)
            parser.set("SOURCE","col_name_min", col_name_min)
            parser.set("SOURCE","col_name_max", col_name_max)
            parser.set("SOURCE","col_name_avg", col_name_avg)
            parser.set("SOURCE","col_name_sum", col_name_sum)
           
            with open("C:\\rulengine_master\configuration.ini", 'w') as file:
                parser.write(file)
            
        if data_source_type == 'JSON':
            file_path1 = request.form['DataSourcePath1'] 
            parser.add_section("APP")
            parser.set("APP",'SOURCE_TYPE',data_source_type)
            parser.set("APP",'OUTPUT_FILE',output_file_path)
            parser.add_section("SOURCE")
            parser.set("SOURCE","SOURCE_DATA_FILE_PATH", file_path1)
            
            with open("C:\\rulengine_master\configuration.ini", 'w') as file:
                parser.write(file)

        if data_source_type == 'XLSX':
            file_path1 = request.form['DataSourcePath1'] 
            sheet_no1 = request.form['sheet_no1'] 
            skip_rows1 = request.form['skip_rows1']
            col_name_min = request.form['min_col']
            col_name_max = request.form['max_col']
            col_name_avg = request.form['avg_col']
            col_name_sum = request.form['sum_col'] 


            parser.add_section("APP")
            parser.set("APP",'SOURCE_TYPE',data_source_type)
            parser.set("APP",'OUTPUT_FILE',output_file_path)

            parser.add_section("SOURCE")
            parser.set("SOURCE","SOURCE_DATA_FILE_PATH", file_path1)
            parser.set("SOURCE","SHEET_NO", sheet_no1)
            parser.set("SOURCE","SKIP_ROWS", skip_rows1)
            parser.set("SOURCE","col_name_min", col_name_min)
            parser.set("SOURCE","col_name_max", col_name_max)
            parser.set("SOURCE","col_name_avg", col_name_avg)
            parser.set("SOURCE","col_name_sum", col_name_sum)

            with open("C:\\rulengine_master\configuration.ini", 'w') as file:
                parser.write(file)

        if data_source_type == 'ORACLE' or data_source_type == 'MYSQL':
            server1 = request.form['Server1'] 
            database1 = request.form['Database1'] 
            user1 = request.form['user1'] 
            password1 = request.form['password1'] 
            schema_name1 = request.form['schema_name1']            
            source_query_filter1 = request.form['source_query_filter1']
            col_name_min = request.form['min_col']
            col_name_max = request.form['max_col']
            col_name_avg = request.form['avg_col']
            col_name_sum = request.form['sum_col']

            parser.add_section("APP")
            parser.set("APP",'SOURCE_TYPE',data_source_type)
            parser.set("APP",'OUTPUT_FILE',output_file_path)

            parser.add_section("SOURCE")
            parser.set("SOURCE","SOURCE_SERVER", server1)
            parser.set("SOURCE","SOURCE_DATABASE", database1)
            parser.set("SOURCE","SOURCE_USER", user1)
            parser.set("SOURCE","SOURCE_PASSWORD", password1)
            parser.set("SOURCE","col_name_min", col_name_min)
            parser.set("SOURCE","col_name_max", col_name_max)
            parser.set("SOURCE","col_name_avg", col_name_avg)
            parser.set("SOURCE","col_name_sum", col_name_sum)

            parser.add_section("vTurbineMasterData_Source")
            parser.set("vTurbineMasterData_Source","SCHEMA_NAME_SOURCE", schema_name1)
            parser.set("vTurbineMasterData_Source","SOURCE_QUERY_FILTER",source_query_filter1)

            with open("C:\\rulengine_master\configuration.ini", 'w') as file:
                parser.write(file)
        
              
        ###################### destination data set User #####################################
                

        data_source_type = request.form['datadesttype']
        parser.set("APP",'DEST_TYPE',data_source_type)

        if data_source_type == 'CSV':

            file2 = request.files['datasourcepath2'] 
            filename2=secure_filename(file2.filename)
            file_path2=os.path.join(basedir + '\Store_File\\' + file2.filename)
            delimiter1 = request.form['Delimiter1']
            skip_rows2 = request.form['skip_rows2']
            delimiter2 = request.form['delimiter2']         
                        
            parser.add_section("DEST")
            parser.set("DEST","DEST_DATA_FILE_PATH", file_path2)
            parser.set("DEST","Delimiter", delimiter2)
            parser.set("DEST","skip_rows", skip_rows2)
            with open("C:\\rulengine_master\configuration.ini", 'w') as file:
                parser.write(file)
           
        if data_source_type == 'JSON':
            file_path2 = request.form['datasourcepath2'] 

            parser.add_section("DEST")
            parser.set("DEST","DEST_DATA_FILE_PATH", file_path2)
         
            with open("C:\\rulengine_master\configuration.ini", 'w') as file:
                parser.write(file)


        if data_source_type == 'XLSX':
            file_path2 = request.form['DataSourcePath2'] 
            sheet_no2 = request.form['sheet_no2'] 
            skip_rows2 = request.form['skip_rows2']

            parser.add_section("DEST")
            parser.set("DEST","DEST_DATA_FILE_PATH", file_path2)
            parser.set("DEST","SHEET_NO", sheet_no2)
            parser.set("DEST","SKIP_ROWS", skip_rows2)

            with open("C:\\rulengine_master\configuration.ini", 'w') as file:
                parser.write(file)


        if data_source_type == 'ORACLE' or data_source_type == 'MYSQL':
            server2 = request.form['Server2'] 
            database2 = request.form['Database2'] 
            user2 = request.form['user2'] 
            password2 = request.form['password2'] 
            schema_name2 = request.form['schema_name2']            
            source_query_filter2 = request.form['source_query_filter2'] 
        
            parser.add_section("DEST")
            parser.set("DEST","DEST_SERVER", server2)
            parser.set("DEST","DEST_DATABASE", database2)
            parser.set("DEST","DEST_USER", user2)
            parser.set("DEST","DEST_PASSWORD", password2)

            parser.add_section("vTurbineMasterData_Dest")
            parser.set("vTurbineMasterData_Dest","SCHEMA_NAME_DEST", schema_name2)
            parser.set("vTurbineMasterData_Dest","Destination_QUERY_FILTER",source_query_filter2)

            with open("C:\\rulengine_master\configuration.ini", 'w') as file:
                parser.write(file)
        
        
        return render_template('UserDoubleDataSource.html')
    except:
        print(Exception)
        raise
        


#app run
if (__name__ == "__main__"):
     app.run(debug=True)

